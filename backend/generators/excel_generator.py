"""
Excel Generator
===============
Uses openpyxl to produce a multi-sheet Excel workbook:
  Sheet 1 : Executive Summary + Key Findings
  Sheet 2+: One sheet per data table found in structured_data
  Last    : Full source bibliography

Styling
-------
- Named styles for headers, body, citations
- Auto column-width adjustment
- Frozen top row on every sheet
"""

import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour definitions (hex, no '#')
# ---------------------------------------------------------------------------
_DARK_BLUE = "1A3A6E"
_MID_BLUE = "3477BD"
_LIGHT_BLUE = "D6E4F7"
_GOLD = "FFBA00"
_LIGHT_GRAY = "F5F6F8"
_WHITE = "FFFFFF"
_TEXT = "212529"

# ---------------------------------------------------------------------------
# Style factories
# ---------------------------------------------------------------------------


def _header_font(bold: bool = True, size: int = 11, color: str = _WHITE) -> Font:
    return Font(bold=bold, size=size, color=color, name="Calibri")


def _body_font(size: int = 10, bold: bool = False, color: str = _TEXT) -> Font:
    return Font(size=size, bold=bold, color=color, name="Calibri")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _wrap_align(h: str = "left", v: str = "center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _set_col_width(ws, col_idx: int, text: str, min_w: int = 12, max_w: int = 60) -> None:
    length = max(min_w, min(len(str(text)) + 4, max_w))
    ws.column_dimensions[get_column_letter(col_idx)].width = length


def _header_row(ws, values: list[str], row: int = 1, fill_color: str = _DARK_BLUE) -> None:
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = _header_font()
        cell.fill = _fill(fill_color)
        cell.alignment = _wrap_align("center")
        cell.border = _border()
        _set_col_width(ws, ci, val)


def _data_cell(ws, row: int, col: int, value, even_row: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _body_font()
    cell.fill = _fill(_LIGHT_GRAY if even_row else _WHITE)
    cell.alignment = _wrap_align()
    cell.border = _border()


def _sheet_summary(wb: openpyxl.Workbook, structured_data: dict, language: str) -> None:
    label = "Resumen" if language == "es" else "Summary"
    ws = wb.active
    ws.title = label

    # Title block
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = structured_data.get("topic", label)
    title_cell.font = Font(bold=True, size=16, color=_WHITE, name="Calibri")
    title_cell.fill = _fill(_DARK_BLUE)
    title_cell.alignment = _wrap_align("center")
    ws.row_dimensions[1].height = 30

    # Executive summary
    row = 3
    label_sum = "Resumen Ejecutivo" if language == "es" else "Executive Summary"
    ws.merge_cells(f"A{row}:D{row}")
    h = ws.cell(row=row, column=1, value=label_sum)
    h.font = _header_font(size=12)
    h.fill = _fill(_MID_BLUE)
    h.alignment = _wrap_align("center")
    row += 1

    ws.merge_cells(f"A{row}:D{row + 4}")
    body = ws.cell(row=row, column=1, value=structured_data.get("executive_summary", ""))
    body.font = _body_font(size=11)
    body.alignment = _wrap_align()
    ws.row_dimensions[row].height = 90
    row += 5

    # Key findings
    row += 1
    label_kf = "Hallazgos Clave" if language == "es" else "Key Findings"
    label_cit = "Cita APA" if language == "es" else "APA Citation"
    _header_row(ws, ["#", label_kf, label_cit, "URL"], row=row, fill_color=_MID_BLUE)
    ws.freeze_panes = ws[f"A{row + 1}"]
    row += 1

    for i, f in enumerate(structured_data.get("key_findings", []), 1):
        even = i % 2 == 0
        _data_cell(ws, row, 1, i, even)
        _data_cell(ws, row, 2, f.get("finding", ""), even)
        _data_cell(ws, row, 3, f.get("citation_apa", ""), even)
        _data_cell(ws, row, 4, f.get("url", ""), even)
        ws.row_dimensions[row].height = 45
        row += 1

    # Auto-width for key columns
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 40


def _sheet_table(wb: openpyxl.Workbook, table_data: dict, language: str) -> None:
    headers = table_data.get("headers", [])
    rows = table_data.get("rows", [])
    title = table_data.get("title", "Table")
    if not headers or not rows:
        return

    # Sanitise sheet name (max 31 chars, no special chars)
    import re
    safe_name = re.sub(r"[\\/*?:\[\]]", "", title)[:31]
    ws = wb.create_sheet(title=safe_name)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=13, color=_WHITE, name="Calibri")
    t.fill = _fill(_DARK_BLUE)
    t.alignment = _wrap_align("center")
    ws.row_dimensions[1].height = 25

    _header_row(ws, headers, row=2, fill_color=_MID_BLUE)
    ws.freeze_panes = ws["A3"]

    for ri, row in enumerate(rows, 3):
        even = ri % 2 == 0
        for ci, val in enumerate(row[:len(headers)], 1):
            _data_cell(ws, ri, ci, val, even)
            _set_col_width(ws, ci, str(val))


def _sheet_sources(wb: openpyxl.Workbook, sources: list[dict], language: str) -> None:
    label = "Bibliografía" if language == "es" else "Bibliography"
    ws = wb.create_sheet(title=label)

    label_title = "Título" if language == "es" else "Title"
    label_apa = "Cita APA" if language == "es" else "APA Citation"
    label_ieee = "Cita IEEE" if language == "es" else "IEEE Citation"
    _header_row(ws, ["#", label_title, label_apa, label_ieee, "URL"], row=1)
    ws.freeze_panes = ws["A2"]

    for i, src in enumerate(sources, 1):
        even = i % 2 == 0
        _data_cell(ws, i + 1, 1, i, even)
        _data_cell(ws, i + 1, 2, src.get("title", ""), even)
        _data_cell(ws, i + 1, 3, src.get("citation_apa", ""), even)
        _data_cell(ws, i + 1, 4, src.get("citation_ieee", ""), even)
        _data_cell(ws, i + 1, 5, src.get("url", ""), even)
        ws.row_dimensions[i + 1].height = 40

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 40


# ---------------------------------------------------------------------------
# Public builder function
# ---------------------------------------------------------------------------


def build_excel(
    topic: str,
    language: str,
    structured_data: dict,
    sources: list[dict],
    extra_instructions: str,
    output_dir: Path,
) -> Path:
    """
    Build an Excel workbook from structured research data.
    Returns path to the saved .xlsx file.
    """
    wb = openpyxl.Workbook()

    _sheet_summary(wb, structured_data, language)

    for table in structured_data.get("data_tables", []):
        _sheet_table(wb, table, language)

    if sources:
        _sheet_sources(wb, sources, language)

    filename = output_dir / f"{uuid.uuid4().hex}.xlsx"
    wb.save(str(filename))
    return filename
